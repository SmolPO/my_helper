from PyQt5 import uic
from PyQt5.QtWidgets import QDialog
import openpyxl as xlxs
from openpyxl.styles import Border, Side, NamedStyle, Font
from openpyxl.styles.borders import BORDER_THIN
import os
from database import *
idx_table = 25


class CreateReport(QDialog):
    def __init__(self, parent):
        super(CreateReport, self).__init__()
        self.conf = Ini(self)
        self.ui_file = self.conf.get_ui("create_report")
        if not self.check_start():
            return
        self.parent = parent
        self.b_ok.clicked.connect(self.ev_ok)
        self.count_part = 0
        self.list_table = []
        self.list_point = []
        self.list_work = []
        self.path = self.conf.get_path("path") + self.conf.get_path("pat_patterns") + "/Исполнительная.xlsx"
        self.data = dict()
        self.list_ui = {"KS2": [
            self.act_1, self.act_2, self.act_3, self.act_4, self.act_5,
            self.act_6, self.act_7, self.act_8, self.act_9, self.act_10],
            "mat": [
                self.mat_1, self.mat_2, self.mat_3, self.mat_4],
            "cult": [
                self.cult_1, self.cult_2, self.cult_3, self.cult_4],
            "table": [
                self.tab_1, self.tab_2],
            "M29": [
                self.stor_1, self.stor_2, self.stor_3, self.stor_4]}
        self.init_bosses()
        self.init_company()
        self.init_customer()
        self.doc = xlxs.open(self.path)
        self.init_dict_cells()
        self.init_dicts()

    # inits
    def init_styles(self):
        ns = NamedStyle(name='thick')
        ns.font = Font(bold=True, size=20)
        border = Side(style='thick', color='000000')
        ns.border = Border(left=border, top=border, right=border, bottom=border)
        self.doc.add_named_style(ns)

        ns = NamedStyle(name='empty')
        ns.font = Font(bold=True, size=20)
        border = Side(style='thin', color='000000')
        ns.border = Border(left=border, top=border, right=border, bottom=border)
        self.doc.add_named_style(ns)

    def init_customer(self):
        rows = self.parent.parent.db.get_data("big_boss, mng_boss, status", "company")
        for row in rows:
            if row[-1] == "Заказчик":
                self.cb_customer.addItems(row[:2])
        pass

    def init_data(self):
        fields = ["number", "date", "object", "type_work", "place", "price"]
        for key in fields:
            self.set_field("contract", key, self.contract[key])
        fields = ["company", "adr", "ogrn", "inn", "kpp", "bik", "korbill", "rbill", "bank",
                  "big_boss", "big_post", "big_at", "big_d_at",
                  "mng_boss", "mng_post", "mng_at", "mng_d_at",
                  "status"]
        for key in fields:
            self.set_field("company", key, self.company[key])
            self.set_field("customer", key, self.customer[key])

    def init_company(self):
        rows = self.parent.parent.db.get_data("*", "company")
        if not rows:
            return False
        for row in rows:
            if row[-2] == "Заказчик":
                self.cb_customer.addItem(row[-1] + ". " + short_name(row[9:12]))
        return rows
        pass

    def init_cult(self):
        part = "culture"
        self.set_field(part, "company", self.company["company"])
        self.set_field(part, "object", self.contract["object"])
        self.set_field(part, "post_1", self.get_post(self.list_ui["cult"][0], "bosses"))
        self.set_field(part, "post_2", self.get_post(self.list_ui["cult"][1], "bosses"))
        self.set_field(part, "post_3", self.get_post(self.list_ui["cult"][2], "bosses"))
        self.set_field(part, "boss_1", self.get_short_name(self.list_ui["cult"][0].currentText()))
        self.set_field(part, "boss_2", self.get_short_name(self.list_ui["cult"][1].currentText()))
        self.set_field(part, "boss_3", self.get_short_name(self.list_ui["cult"][3].currentText()))

    def init_ks2(self):
        part = "KS2"
        self.set_field(part, "title", "сдачи-приемки выполненных  № 1")
        self.set_field(part, "city", self.parent.parent.city)
        self.set_field(part, "object", self.contract["object"])
        self.set_field(part, "type_work", self.contract["type_work"])
        note = 'ПАО "Дорогобуж", ОГРН 1026700535773, именуемый в дальнейшем "Заказчик" в лице ' \
               'Заместителя исполнительного директора - Главного инженера Симакова Павла Николаевича, ' \
               'действующего на основании доверенности № 04-Д/34 от 26.05.2021 г. с одной стороны, и ' \
               'ООО "Вертикаль", ОГРН 11967330185475, именуемое в дальнейшем "Подрядчик" ' \
               'в лице генерального директора Тимашева Юрия Николаевича, действующего на основании "Устава", ' \
               'с другой стороны, составили настоящий Акт сдачи-приемки выполненных работ, ' \
               'именуемый в дальнейшем "Акт", о нижеследующем:'
        note = '{0}, ОГРН {1}, именуемый в дальнейшем "Заказчик" в лице ' \
               '{2} {3}, действующего на основании {4}. с одной стороны, и ' \
               '{5}, ОГРН {6}, именуемое в дальнейшем "Подрядчик" ' \
               'в лице {7} {8}, действующего на основании {9}, ' \
               'с другой стороны, составили настоящий Акт сдачи-приемки выполненных работ, ' \
               'именуемый в дальнейшем "Акт", о нижеследующем:'.format(self.customer["company"], self.customer["ogrn"],
                                                                       self.customer["big_post"], self.customer["big_boss"],
                                                                       self.customer["big_at"],
                                                                       self.company["company"], self.company["ogrn"],
                                                                       self.company["big_post"], self.company["big_boss"],
                                                                       self.company["big_at"])
        self.set_field(part, "note", note)
        self.init_table_ks2()

    def init_mat(self):
        part = "mat"
        self.set_field(part, "title", "Ведомость материалов к акту № 1")

    def init_productions(self):
        part = "production"
        self.set_field(part, "part", self.contract["part"])
        self.set_field(part, "object", self.contract["object"])
        self.set_field(part, "contract", self.contract["contract"])
        self.set_field(part, "company", self.company["name"])
        self.set_field(part, "post_boss", "")
        self.set_field(part, "boss", "")
        self.set_field(part, "title", "Расчет выработки за ноябрь месяц, приложение к акту № 1")

    def init_table(self):
        part = "table"
        self.set_field(part, "object", self.contract["object"])
        self.set_field(part, "object", "ноябрь" + str(dt.datetime.now().year) + "г.")

    def init_m29(self):
        part = "M29"
        self.set_field(part, "contract", "Договор подряда № " + self.contract["number"] + " от " + self.contract["date"])
        self.set_field(part, "add", "Приложение №2 к договору" + self.contract["number"] + " от " + self.contract["date"])
        self.set_field(part, "post_1", self.get_post(self.list_ui["M29"][0], "bosses"))
        self.set_field(part, "post_2", self.get_post(self.list_ui["M29"][1], "bosses"))
        self.set_field(part, "post_3", self.get_post(self.list_ui["M29"][2], "bosses"))
        self.set_field(part, "boss_1", self.get_short_name(self.list_ui["M29"][0].currentText()))
        self.set_field(part, "boss_2", self.get_short_name(self.list_ui["M29"][1].currentText()))
        self.set_field(part, "boss_3", self.get_short_name(self.list_ui["M29"][3].currentText()))

    def init_table_ks2(self):
        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00000000'),
            right=Side(border_style=BORDER_THIN, color='00000000'),
            top=Side(border_style=BORDER_THIN, color='00000000'),
            bottom=Side(border_style=BORDER_THIN, color='00000000')
        )
        parts = self.count.value()
        points = self.points.toPlainText().split(",")
        rows = self.rows.toPlainText().split(",")
        count_rows = parts
        for item in rows:
            count_rows += int(item)
        for row in range(14, 14 + count_rows):
            for column in ["A", "B", "C", "D", "E", "F"]:
                cell = column + str(row)
                self.sheet[cell].border = thin_border
            pass
        cursor = 14
        prev = 0
        for ind in range(len(points)):
            self.list_table.append(rows[prev:prev+int(points[ind])])
            prev = int(points[ind])
        for part in self.list_table:
            cursor += 1
            for point in range(len(part)):
                text = self.create_point(point, int(part[point]))
                for i in range(int(part[point])):
                    row = cursor
                    self.set_text(row, text[i])
                    if part[point] == '1':
                        self.list_work.append(cursor)
                        continue
                    if text[i][1] == "работа":
                        self.list_work.append(cursor)
                        continue
                    if i == 0:
                        self.list_point.append("F" + str(cursor))
                    cursor += 1
        self.create_result(cursor)
        self.create_footer(cursor)

    def init_ks3(self):
        self.set_field("KS3", "investor", "Инвестор ")
        self.set_field("KS3", "company", "Подрядчик")
        self.set_field("KS3", "customer", "Заказчик")
        self.set_field("KS3", "contract", self.contract["number"])
        self.set_field("KS3", "day", self.contract["date"][:2])
        self.set_field("KS3", "month", self.contract["date"][3:5])
        self.set_field("KS3", "year", self.contract["date"][6:10])
        self.set_field("KS3", "make_date", "_")
        date = "01." + str(dt.datetime.now().month) + "." + str(dt.datetime.now().year)
        end_date = str(count_days[dt.datetime.now().month]) + str(dt.datetime.now().month) + "." + str(dt.datetime.now().year)
        self.set_field("KS3", "start_date", date)
        self.set_field("KS3", "end_date", end_date)
        self.set_field("KS3", "price_month_1", self.cells["KS2"][1]["sum"])
        self.set_field("KS3", "price_month_2", self.cells["KS2"][1]["sum"])
        self.set_field("KS3", "price_month_3", self.cells["KS2"][1]["sum"])
        self.set_field("KS3", "nds", int(self.cells["KS2"][1]["sum"]) * 0.2)
        self.set_field("KS3", "price", int(self.cells["KS3"][1]["sum"]) * 1.2)

        pass

    def create_footer(self, cursor):
        pass

    def create_prod(self):
        sum = 0
        for cell in self.list_point:
            sum += int(self.sheet.cell(cell).value)
        self.set_field("culture", "sum", sum)

    def create_result(self, cursor):
        cells = "A" + str(cursor) + ":E" + str(cursor)
        self.sheet.merge_cells(cells)
        self.sheet[cells[:2]].value = "Итого по разделам:"
        self.sheet["F" + str(cursor)].value = "=" + "+".join(self.list_point)
        cursor += 1
        cells = "A" + str(cursor) + ":E" + str(cursor)
        self.sheet.merge_cells(cells)
        self.sheet[cells[:2]].value = "НДС 20%:"
        self.sheet["F" + str(cursor)].value = "=" + "F" + str(cursor-1) + "*0,2"
        cursor += 1
        cells = "A" + str(cursor) + ":E" + str(cursor)
        self.sheet.merge_cells(cells)
        self.sheet[cells[:2]].value = "Всего с НДС:"
        self.sheet["F" + str(cursor)].value = "=" + "F" + str(cursor-1) + "F" + str(cursor-2)
        self.cells["KS2"][1]["sum"] = "F" + str(cursor)

    def get_cells(self, cell):
        return "A" + str(cell) + ":F" + str(cell)

    def set_text(self, row, text):
        i = 0
        for column in ["A", "B", "C"]:
            cell = column + str(row)
            self.sheet[cell].value = text[i]
            i += 1

    def create_row(self, row, style):
        for column in ["A", "B", "C", "D", "E", "F"]:
            cell = column + str(row)
            self.sheet[cell].borders = style
        pass

    def create_part(self, _row):
        row = str(_row)
        cells = "A" + row + ":F" + row
        self.sheet.merge_cells(cells)
        self.sheet["A" + row].style = "thick"
        pass

    def create_point(self, point, count_rows):
        if count_rows == 1:
            return [(str(point + 1), "", "")]
        if count_rows == 2:
            return [(str(point + 1), "", ""),
                    (str(point + 1) + ".1", "работа", "")]
        if count_rows < 5:
            return [(str(point + 1), "", ""),
                    (str(point + 1) + ".1", "работа", ""),
                    ("", "", "") * (count_rows-2)]
        if count_rows >= 5:
            return [(str(point + 1), "", ""),
                    (str(point + 1) + ".1", "работа", ""),
                    (str(point + 1) + ".2", "Материал Заказчика", "руб."),
                    ("", "Вспомогательные материалы", ""),
                    (str(point + 1) + ".3", "Материалы Подрядчика", "руб."),
                    ("", "", "") * (count_rows-5)]

    def get_dict(self, table):
        _fields = {"itrs": "family, name, surname, post, status, id",
                   "contracts": "name, customer, number, date, object, type_work, "
                                "place, price, date_end, nds, status, id",
                   "company": "company, adr, ogrn, inn, kpp, bik, korbill, rbill, bank, "
                              "big_boss, big_post, big_at, big_d_at, "
                              "mng_boss, mng_post, mng_at, mng_d_at, "
                              "status, id",
                   "bosses": "family, name, surname, post, email, phone, sex, status, id"}
        fields = _fields.get(table)
        data = list()
        rows = self.parent.parent.db.get_data(fields, table)
        for ind in range(len(rows)):
            j = iter(range(len(fields.split(", "))))
            row = dict()
            for key in fields.split(", "):
                row[key] = rows[ind][next(j)]
            data.append(row)
        return data

    def get_post(self, my_id, table):
        rows = self.parent.parent.db.get_data("*", table)
        for item in rows:
            if str(item[-1]) == my_id:
                print(item)
                return item[3]
        return "."

    def get_short_name(self, text):
        return "".join(text.split(". ")[1:])

    def get_data(self):
        data = dict()
        return data

    def ev_ok(self):
        self.create_report()
        self.close()

    def create_report(self):
        path = self.conf.get_path("path") + self.conf.get_path("contracts") + "/1030/result.xlsx"

        self.init_data()
        self.doc.save(path)
        self.doc.close()
        os.startfile(path)
        self.doc = xlxs.open(self.path)

        self.init_ks2()
        self.doc.save(path)
        self.doc.close()
        os.startfile(path)
        self.doc = xlxs.open(self.path)

        self.init_m29()
        self.doc.save(path)
        self.doc.close()
        os.startfile(path)
        self.doc = xlxs.open(self.path)

        self.init_mat()
        self.doc.save(path)
        self.doc.close()
        os.startfile(path)
        self.doc = xlxs.open(self.path)

        self.init_table()
        self.doc.save(path)
        self.doc.close()
        os.startfile(path)
        self.doc = xlxs.open(self.path)

        self.init_cult()
        self.doc.save(path)
        self.doc.close()
        os.startfile(path)
        self.doc = xlxs.open(self.path)

    def init_dict_cells(self):
        self.cells = {"contract": ("result", {
                          "number": "B2",
                          "date": "B3",
                          "object": "B4",
                          "type_work": "B5",
                          "place": "B6",
                          "price": "B7",
                          "did": "B8"}),
                      "company": ("result", {
                          "company": "B11",
                          "big_boss": "B12",
                          "big_post": "",
                          "big_at": "",
                          "big_d_at": "",
                          "mng_boss": "B12",
                          "mng_post": "",
                          "mng_at": "",
                          "mng_d_at": "",
                          "inn": "B14",
                          "bik": "B15",
                          "kpp": "B16",
                          "ogrn": "B17",
                          "korbill": "B18",
                          "rbill": "B19",
                          "bank": "B20",
                          "adr": "B21"}),
                      "customer": ("result", {
                          "company": "B23",
                          "big_boss": "B24",
                          "inn": "B25",
                          "bik": "B26",
                          "kpp": "B27",
                          "ogrn": "B28",
                          "korbill": "B29",
                          "rbill": "B30",
                          "bank": "B31",
                          "adr": "B32"}),
                      "KS2": ("КС2", {
                          "title": "A2",
                          "note": "A6",
                          "city": "A4",
                          "p1": "A7",
                          "object": "A10",
                          "type_work": "A13",
                          "row": "A16",
                          "sum": ""}),
                      "culture": ("КУЛЬТУРА", {
                          "company": "D16",
                          "object": "A12",
                          "post_1": "A34",
                          "post_2": "A37",
                          "post_3": "A40",
                          "boss_1": "J34",
                          "boss_2": "J37",
                          "boss_3": "J40"}),
                      "production": ("ВЫРАБОТКА", {
                          "part": "B7",
                          "object": "B8",
                          "contract": "C7",
                          "price": "D7",
                          "prod": "E7",
                          "days": "F7",
                          "prod_day": "G7",
                          "prod_month": "H7",
                          "company": "I7",
                          "title": "A1",
                          "post_boss": "B11",
                          "boss": "D11"}),
                      "table": ("ТАБЕЛЬ", {
                          "object": "A2",
                          "date": "A3",
                          "worker": "A6"}),
                      "M29": ("М29", {
                          "company": "C2",
                          "big_boss_1": "B2",
                          "part": "B5",
                          "type_work": "B6",
                          "object": "B7",
                          "contract": "B8",
                          "add": "B9",
                          "date": "J21",
                          "big_boss_2": "C26",
                          "sing_boss": "E30",
                          "post_1": "F44",
                          "post_2": "F46",
                          "boss_1": "N44",
                          "boss_2": "N46",
                          "point": "B53",
                          "post_3": "",
                          "boss_3": "",
                          "date_2": ""}),
                      "KS3": ("КС3", {
                          "investor": "",
                          "company": "",
                          "customer": "",
                          "build": "",
                          "contract": "",
                          "day": "",
                          "month": "",
                          "year": "",
                          "count": "",
                          "make_date": "",
                          "start_date": "",
                          "end_date": "",
                          "work": "",
                          "price_all": "",
                          "price_year": "",
                          "price_month_1": "",
                          "price_month_2": "",
                          "price_month_3": "",
                          "nds": "",
                          "price": "",
                          "post_1": "",
                          "boss_1": ""}),
                      "mat": ("МАТЕРИАЛЫ", {
                          "title": "A1",
                          "material": "A6"})
                      }

    def set_val(self, sheet, cell, val):
        self.sheet = self.doc[sheet]
        self.sheet.cell(cell).value = val
        pass

    def set_field(self, part, field, val):
        self.sheet = self.doc[self.cells.get(part)[0]]
        cell = self.cells.get(part)[1].get(field)
        print(cell, part, field, val)
        if cell == "" or not cell:
            print("NOT ADD")
            return
        self.sheet[cell].value = val
        pass

    def add_val(self, row, column, val):
        self.sheet.cell(row=row, column=column).value = val

    def create_ekr(self, data, sheet):
        """
        1;3;7
        :param data:
        :return:
        """
        parts = data.split(";")  # 5;5;8
        for item in parts:
            sheet.insert_rows(idx=idx_table, amount=5*item)
            cells = sheet['A'+str(idx_table):'F'+5*item]
            for cell in cells:
                cell.style.borders.left.border_style = Border.BORDER_THIN
                if int(cells.coordinate[1:])-idx_table/5 == 0:
                    cell.value = "работа"

            sheet.insert_rows(idx=idx_table)
            sheet.merge_cells(start_row=idx_table, start_column=1, end_row=idx_table, end_column=6)
            sheet.cell(row=idx_table, colomn=1).style.borders.left.border_style = Border.BORDER_THICK
        pass

    def check_input(self):
         return True

    def init_bosses(self):
        for key in self.list_ui.keys():
            for ui in self.list_ui.get(key):
                ui.addItem("(нет)")
                self.parent.parent.parent.db.init_list(ui, "*", "bosses", people=True)
                self.parent.parent.parent.db.init_list(ui, "*", "itrs", people=True)
        pass

    def init_dicts(self):
        contracts = self.get_dict("contracts")
        company = self.get_dict("company")
        self.bosses = self.get_dict("bosses")
        self.itrs = self.get_dict("itrs")
        self.number = get_val(self.parent.cb_select)
        for ind in range(len(contracts)):
            if self.parent.cb_select.currentText().split(". ")[0] == contracts[ind]["id"]:
                self.contract = contracts[ind]
        for ind in range(len(company)):
            if self.parent.parent.company == company[ind]["company"]:
                self.company = company[ind]
                self.customer = company[ind]
            print(self.parent.cb_comp.currentText().split(". ")[0])
            if self.parent.cb_comp.currentText().split(". ")[0] == company[ind]["id"]:
                self.customer = company[ind]
        pass

    def check_start(self):
        self.status_ = True

        try:
            uic.loadUi(self.ui_file, self)
            return True
        except:
            mes.question(self, "Сообщение", "Не удалось открыть форму " + self.ui_file, mes.Cancel)
            self.status_ = False
            return False