import openpyxl as xlsx
from openpyxl.styles import Side, Border
from database import *


class NewCovid:
    def __init__(self, parent=None):
        self.conf = Ini(self)
        self.db = DataBase(parent)
        self.parent = parent
        self.border = Border(top=Side(border_style='thin', color='FF000000'),
                           right=Side(border_style='thin', color='FF000000'),
                           bottom=Side(border_style='thin', color='FF000000'),
                           left=Side(border_style='thin', color='FF000000'))

    def create_covid(self):
        list_people = list()
        rows_ = [self.db.get_data("family, name, surname, post, status, id", WORKERS),
               self.db.get_data("family, name, surname, post, status, id", ITRS)]
        if ERR in rows_:
            return
        rows = rows_[0] + rows_[1]
        if not rows:
            return msg_info(self, GET_DB)
        for row in rows:
            if row[-2] == statues[0]:
                list_people.append({"family": short_name(row), "post": row[3]})
        list_people.sort(key=lambda x: x["family"])
        if not list_people:
            mes.question(self.parent, "Сообщение", "Нет рабочих в Базе данных", mes.Cancel)
            return
        path = self.conf.get_path("pat_covid")
        path_save = path
        try:
            doc = xlsx.open(path)
        except:
            return msg_er(self.parent, GET_FILE + path)
        page = "covid"
        try:
            sheet = doc[page]
        except:
            return msg_er(self.parent, GET_PAGE + page)
        delta = 2
        count_column = 9
        for ind in range(1, 50):
            for j in range(1, 10):
                sheet.cell(row=ind + delta, column=j).value = ""
        for ind in range(1, len(list_people) + 1):
            sheet.cell(row=ind + delta, column=1).value = ind
            sheet.cell(row=ind + delta, column=2).value = dt.datetime.now().date()
            sheet.cell(row=ind + delta, column=3).value = list_people[ind-1]["family"]
            sheet.cell(row=ind + delta, column=4).value = list_people[ind-1]["post"]
            for i in range(1, 9):
                sheet.cell(row=ind + delta, column=count_column).border = self.border
        doc.print_area = "A1:G" + str(len(list_people) + delta)
        try:
            doc.save(path_save)
            print_to(path_save)
        except:
            return msg_er(self.parent, GET_FILE + path_save)
